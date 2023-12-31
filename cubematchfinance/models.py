import os
import re

from PyQt6.QtWidgets import QFileDialog, QMessageBox
from PyQt6.QtCore import QStandardPaths

import PyPDF2
import pdfplumber
import pandas as pd
import subprocess
import openpyxl
import numpy as np
from datetime import datetime

class Model:

    def __init__(self):
       
        self.tab1 = self.Tab1(self)
        self.tab2 = self.Tab2(self)
        self.tab3 = self.Tab3(self)
        self.tab4 = self.Tab4(self)
        self.tab5 = self.Tab5(self)
        self.tab6 = self.Tab6(self)
        self.tab7 = self.Tab7(self)

    def sanitize_filename(self, filename):

        return re.sub(r'[<>:"/\\|?*]', '', filename)

    def rename(self, old_path, new_path):

        if os.path.exists(new_path):
            message = f"{new_path} Already exist"
            QMessageBox.critical(None, "Error", message)
        else:
            try:
                os.rename(src=old_path, dst=new_path)
            except Exception as e:
                message = f"An error occurred while renaming file: {str(e)}"
                QMessageBox.critical(None, "Error", message)
    
    class Tab1:

        def __init__(self, parent):

            self.parent = parent
        
        def browse(self, button):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def renaming_timesheets(self):

            try:
                file_names = self.fileName
                for pdf in file_names:
                    folder_path = os.path.dirname(pdf)
                    if pdf.endswith('.pdf'):
                        old_path = pdf
                        with pdfplumber.open(old_path) as pdf:
                            self.__renaming(pdf, folder_path, old_path)
            except Exception as e:
                    error_message = f"An error occurred while renaming info: {str(e)}"
                    QMessageBox.critical(None, "Error", error_message)

        def __renaming(self, pdf, folder_path, old_path):

            pages = pdf.pages
            pagina = pages[0]
            data = pagina.extract_text_lines()
            new_name = self.__set_name_timesheet(data)
            new_path = os.path.join(folder_path, new_name)
            pdf.close()
            self.parent.rename(new_path, old_path)
            
         
        def __set_name_timesheet(self, data):

            body = data[0]['text']
            name = data[1]['text'].split()[2:4]
            name = ' '.join(name)
            client = data[5]['text'].split()[1:]
            client = ' '.join(client)
            date = pd.to_datetime(data[2]['text'].replace('Timesheet period: ', '')[len('DD/MM/YYYY '):])
            new_name = f'{name} ({client}) {body} - {str(date.month_name())} {date.year}.pdf'
            new_name = self.parent.sanitize_filename(new_name)
            return new_name

    class Tab2:

        def __init__(self, parent):

            self.parent = parent
            self.fileName = []
            self.split_pdf_files = []
        
        def browse(self, button):
            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def split_pdf(self):

            try:
                files = self.fileName[0]
                self.split_pdf_files = []
                folder_path = os.path.dirname(files)
                output_dir = os.path.join(folder_path, 'Sales')
                os.makedirs(output_dir, exist_ok=True)
            
                with open(files, 'rb') as file:
                    pdf = PyPDF2.PdfReader(file)
                    total_pages = len(pdf.pages)
                    for page_number in range(total_pages):
                        pdf_writer = PyPDF2.PdfWriter()
                        pdf_writer.add_page(pdf.pages[page_number])
                        output_file_name = f'page_{page_number + 1}.pdf'
                        output_file_path = os.path.join(output_dir, output_file_name).replace("\\\\", "\\")
                        self.split_pdf_files.append(output_file_path)
                        with open(output_file_path, 'wb') as output_file:
                            pdf_writer.write(output_file)
            except Exception as e:
                error_message = f"An error occurred while splitting files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def renaming_invoices(self):

            try:
                pdf_files = self.split_pdf_files
                folder_path = os.path.dirname(pdf_files[0])
                for file in pdf_files:
                    if file.endswith('.pdf'):
                        old_path = os.path.join(folder_path, file)
                        with pdfplumber.open(old_path) as pdf:
                            page = pdf.pages[0].extract_tables()
                            new_name = self.__set_name_sales_invoices(page)
                            new_path = os.path.join(folder_path, new_name)
                            pdf.close()
                            self.parent.rename(old_path, new_path)
            except Exception as e:
                error_message = f"An error occurred while renaming files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)


        def __set_name_sales_invoices(self, page):

            try:
                invoice_number = page[0][2][1]
                reference_number = page[0][3][1]
                name = page[1][2][0].split()[:2]
                name = ' '.join(name)
                date = page[0][0][1][2:]
                new_name = f"{invoice_number}.{reference_number} ({name}) -{date}.pdf"
                new_name = self.parent.sanitize_filename(new_name)
                return new_name
            except Exception as e:
                error_message = f"An error occurred while extracting info: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
                            
    class Tab3:

        def __init__(self, parent):

            self.parent = parent
        
        def browse(self, button):
            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
    
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
    
        def __docxtopdf(self, file):

            try:
                files_path = os.path.dirname(file)
                nombre_archivo, extension = os.path.splitext(file)
                file_path = os.path.join(files_path, file)
                pdf_path = os.path.join(files_path, f'{nombre_archivo}.pdf')
                with open(os.devnull, 'w') as nullfile:
                    subprocess.run(["docx2pdf", file_path, pdf_path], stdout=nullfile, stderr=nullfile, shell=True)
                os.remove(file_path)
                return pdf_path
            except Exception as e:
                error_message = f"An error occurred while converting to .pdf: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def browse_pb(self):

            try:
                self.pb, _ = QFileDialog.getOpenFileNames(None, 'Open File')
                
            except Exception as e:
                error_message = f"An error occurred  while getting Purchase Book: {str(e)}"
                print(error_message)
        
        def extract_purchasebook(self):

            try:
                self.df_pb = pd.read_excel(self.pb[0])
                return self.df_pb
            except Exception as e:
                error_message = f"No Purchase Book selected: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def renaming_contractors(self, files):

            pb = self.extract_purchasebook()
            files = self.__converter()
            for file in files:
                try:
                    old_path = os.path.dirname(file)
                    full_name = os.path.basename(file)
                    company = full_name.split('_')[0]
                    date = pd.to_datetime(full_name.split('_')[1])
                    company_trim, pb_trim = self.__trim_str(company, pb)
                    
                    if company_trim in pb_trim:
                        try:
                            new_name = self.__find_contractor_name(pb, company, date)
                            new_path = os.path.join(old_path, new_name)
                            os.rename(file, new_path)
                        except:
                            self.__renaming(file, old_path) 
                    else:
                        self.__renaming(file, old_path) 

                except Exception as e:
                    error_message = f"An error occurred while renaming: {str(e)}"
                    QMessageBox.critical(None, "Error", error_message)

        def __converter(self):
            
            try:
                new_file_names = []
                for file in self.fileName:
                    if file.endswith('.docx'):
                        file = self.__docxtopdf(file)
                    new_file_names.append(file)
                return new_file_names
            except:
                AttributeError      

        def __trim_str(self, company, pb):
         
            company_trim = company.lower().replace(" ", "")
            pb_trim = pb.iloc[:, 0].str.lower().str.replace(" ", "").values
            return company_trim, pb_trim

        def __find_contractor_name(self, pb, company, date):
         
            mask = pb.iloc[:, 0].str.lower().str.replace(" ", "") == company.lower().replace(" ", ""),pb.columns[1]
            contractor_name = pb.loc[mask].values[0]
            new_name = f"{company} ({contractor_name}) - {str(date.month_name())} {date.year}.pdf"
            new_name = self.parent.sanitize_filename(new_name)

            return new_name

        def __set_name_non_contractors(self, file_path, existing_files):
            full_name = os.path.basename(file_path)
            company_name = full_name.split('_')[0]
            date_str = full_name.split('_')[1]
            date = pd.to_datetime(date_str)
            formatted_date = f'{date.strftime("%B")} {date.year}'
            new_name = f'{company_name} - {formatted_date}.pdf'
            if new_name in existing_files:
                base_name, extension = os.path.splitext(new_name)
                new_name = f'{base_name} ({self.__get_next_unique_number(existing_files)}){extension}'
            
            return new_name

        def __get_next_unique_number(self, existing_files):

            secuential_number = 1
    
            while True:
                candidate = secuential_number
                if not any(f"_ {candidate}" in file for file in existing_files):
                    return candidate
                secuential_number += 1
        
        def __renaming(self, file, old_path):
            
            existing_files = os.listdir(old_path)
            new_name = self.__set_name_non_contractors(file, existing_files)
            new_path = os.path.join(old_path, new_name)
            os.rename(file, new_path)

        def renaming_noncontractors_other(self, files):
            
            files = self.__converter()
            if files is not None:
                old_path = os.path.dirname(files[0])
                for file in files:
                    try:
                        self.__renaming(file, old_path)
                    except Exception as e:
                        error_message = f"An error occurred while renaming: {str(e)}"
                        QMessageBox.critical(None, "Error", error_message)

    class Tab4:

        def __init__(self, parent):

            self.parent = parent
            
        def browse(self, button):
            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def salaries_extract(self):
            self.information = []
            try:
                file_names = self.fileName
                with pdfplumber.open(file_names[0]) as pdf:
                    page = pdf.pages[0].extract_tables()[0]
                    for fila in page:
                        if isinstance(fila, (list, dict)):
                            if isinstance(fila, dict):
                                if fila.get(0) is not None:
                                    self.information.append(fila)
                            elif isinstance(fila, list):
                                if fila[0].isdigit():
                                    self.information.append(fila)
            except Exception as e:
                error_message =f"An error occurred while extracting PDF info:{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
            return self.information

        def write_excel(self):
           
            try:
                information = self.salaries_extract()
                if information:
                    libro_excel, _ = QFileDialog.getOpenFileNames(None, "Select Files", r"C:\\")
                    archivo_excel = libro_excel[0]
                    libro_excel = openpyxl.load_workbook(archivo_excel)
                    replace_dict = {'.': '', ',': '.'}
                    
                    spreadsheet = libro_excel.active
            
                    for fila, datos in enumerate(information, start=12):
                        spreadsheet.cell(row=fila, column=2).value = datos[1]  
                        spreadsheet.cell(row=fila, column=3).value = datos[2] 
                        spreadsheet.cell(row=fila, column=4).value = datos[3].translate(str.maketrans(replace_dict))
                        spreadsheet.cell(row=fila, column=5).value = "CM Salary"
                    
                    libro_excel.save(archivo_excel)

            except Exception as e:
                error_message =f"An error occurred while writing on Excel: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)

    class Tab5:

        def __init__(self, parent):

            self.parent = parent

        def browse(self, button):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def extract_clarity_details(self):
       
            exceptions = ['Weedle']
            try:
                df = self.clean_clarity()
                df_grouped = df.groupby(['Contractor PO Number', 'Surname', 'First Name']).agg({'Sum of Hours': 'sum', 'Sum of Total': 'sum'}).reset_index()
                mask = df_grouped['Surname'].isin(exceptions)
                df_grouped['Total Days'] = df_grouped['Sum of Hours'] / mask.replace({True: 7.5, False: 8})
                self.df_result = df_grouped
            except Exception as e:
                error_message =f"Error while extracting Clarity information: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
            else:
                return  self.df_result

        
        def clean_clarity(self):
            try:
                file_names = self.fileName[0]
                self.folder_path = os.path.dirname(file_names)
                df = pd.read_excel(file_names, engine='pyxlsb', skiprows=1, index_col=0)
                df.columns = df.iloc[df.index.get_loc('Resource ID')]
                mask = pd.notnull(df.columns)
                df = df.loc [:, mask] 
                df = df[~df.isna().any(axis=1)]
                df = df.iloc[1:]
                return df
            except Exception as e:
                error_message =f"Error while cleaning Clarity information: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def export_clarity_to_excel(self):
            try:
                self.df_result.to_excel(os.path.join(self.folder_path, "Clarity Details.xlsx"), index = False)
            except Exception as e:
                error_message =f"Error while exporting Clarity information: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
    class Tab6:

        def __init__(self, parent):

            self.parent = parent
        
        def browse(self):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(None, 'Open File')
            except Exception as e:

                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def __assingorderbook(self):
            try:
                for name in self.fileName:
                    if "CMIRE" in name:
                        self.IRE = pd.read_excel(name, skiprows=3, skipfooter=7)
                    if "UK" in name:
                        self.UK = pd.read_excel(name, skiprows=3, skipfooter=7)
                    if "NL" in name:
                        self.NL = pd.read_excel(name, skiprows=3, skipfooter=6)
            except:
                error_message = f"An error occurred while obtaining books. Remember that books must not have password."
                QMessageBox.critical(None, "Error", error_message)
                
        def order_book_cleaner(self):
            try:
                self.__assingorderbook()
                self.__first_step()
                self.__second_step()
                self.__third_step()
                self.__fourth_step()
                self.__fifth_step()
                self.__sixth_step()
                self.__seventh_step()
                self.__eighth_step()
                self.__ninth_step()
                self.__tenth_step()
                self.__index_creation()
                return self.OB2023
            except:
                error_message = f"An error occurred while cleaning books"
                QMessageBox.critical(None, "Error", error_message)
        
        def export_order_book(self):
            try:
                folder_path = os.path.dirname(self.fileName[0])
                return self.OB2023.to_excel(os.path.join(folder_path,'OB-ThisYear.xlsx'))
            except:
                error_message = f"An error occurred while exporting book:\n"
                QMessageBox.critical(None, "Error", error_message)
        
        def __first_step(self):
            self.IRE = self.IRE.loc[:, ~self.IRE.columns.str.startswith('Unnamed')]
            self.UK = self.UK.loc[:, ~self.UK.columns.str.startswith('Unnamed')]
            self.NL = self.NL.loc[:, ~self.NL.columns.str.startswith('Unnamed')]
           
    
        def __second_step(self):
            self.IRE.drop(labels=0, axis = 0, inplace= True)
            self.UK.drop(labels=0, axis = 0, inplace= True)
            self.NL.drop(labels=0, axis = 0, inplace= True)
            self.IRE.drop(columns='Comments', inplace=True)
            self.UK.drop(columns='Comments', inplace=True)
            self.NL.drop(columns='Comments', inplace=True)
        
        def __third_step(self):
            columns_to_renameUK = {'Cost ': 'Cost Rate','Sale':'Sale Rate'}
            columns_to_renameBV = {'Pay Rate': 'Cost Rate', 'Sale Rate/Fee':'Sale Rate'}

            self.UK.rename(columns=columns_to_renameUK, inplace=True)
            self.NL.rename(columns=columns_to_renameBV, inplace=True)
            
        
        def __remove_estimations(self, data):
            current_month = datetime.now().strftime(('%B'))
            position_current_month = data.columns.get_loc(current_month)
            data.iloc[:,position_current_month:] = np.nan
            

        def __fourth_step(self):
            self.__remove_estimations(self.IRE)
            self.__remove_estimations(self.UK)
            self.__remove_estimations(self.NL)

        def __fifth_step(self):
            self.IRE['Location'] = 'Ireland'
            self.UK['Location'] = 'United Kingdom'
            self.NL['Location'] = 'Netherlands'
            self.UK['Project'] = np.nan
            self.NL['Project'] = np.nan
            self.NL['Role'] = np.nan
            self.IRE.Project.replace(np.nan, 'Unknown', inplace=True)
            self.IRE.Role.replace(np.nan, 'Unknown', inplace=True)
            self.IRE['CM Role'].replace(np.nan, 'Unknown', inplace=True)
            self.UK.Project.replace(np.nan, 'Unknown', inplace=True)
            self.UK.Role.replace(np.nan, 'Unknown', inplace=True)
            self.NL.Project.replace(np.nan, 'Unknown', inplace=True)
            self.NL.Role.replace(np.nan, 'Unknown', inplace=True)
            

        def __sixth_step(self):
            for column in ['End Date','Start Date']:
                self.IRE[column]= pd.to_datetime(self.IRE[column],format='%d/%m/%Y',)
                self.UK[column] = pd.to_datetime(self.UK[column], format='%d/%m/%Y')
                self.NL[column] = pd.to_datetime(self.NL[column], format='%d/%m/%Y')
            
        
        def __seventh_step(self):
            self.IRE['Cost Rate'] = ['Permanent' if x == 'Perm' else 'Contractor' for x in self.IRE['Cost Rate']]
            self.UK['Cost Rate'] = ['Permanent' if x == 'Perm' else 'Contractor' for x in self.UK['Cost Rate']]
            self.NL['Cost Rate'] = ['Permanent' if x == 'Perm' else 'Contractor' for x in self.NL['Cost Rate']]
        
        def __eighth_step(self):
            cols = self.UK.columns.tolist()
            cols.insert(2, cols.pop(cols.index('Project')))
            self.UK = self.UK[cols]
            cols = self.NL.columns.tolist()
            cols.insert(2, cols.pop(cols.index('Project')))
            cols.insert(3, cols.pop(cols.index('Role')))
            self.NL = self.NL[cols]
            self.NL['Sale Rate'] = self.NL['Sale Rate'] * 8
            for column in self.NL.iloc[:,8:-1].columns:
                self.NL[column].astype(float) 
                self.NL[column] = self.NL[column]/8
        
        def __ninth_step(self):
            self.OB2023 = pd.concat([self.IRE, self.UK, self.NL])
            self.OB2023['Year'] = 2023
            self.OB2023['generation_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        

        def __tenth_step(self):
            branches = ['NL','IRE','UK']
            for branch in branches:
                if branch =='NL':
                    self.OB2023.loc[self.OB2023['Client'].str.contains(branch), 'Client'] = 'CMNL'
                if branch == 'IRE':
                    self.OB2023.loc[self.OB2023['Client'].str.contains(branch), 'Client'] = 'CMIRE'
                if branch == 'UK':
                    self.OB2023.loc[self.OB2023['Client'].str.contains(branch), 'Client'] = 'CMUK'
            self.OB2023['CM Role'].fillna('Unknown', inplace=True)
        
        def __index_creation(self):
            self.OB2023.reset_index(inplace=True)
            self.OB2023.drop(columns='index', inplace=True)

    
    class Tab7:

        def __init__(self, parent):

            self.parent = parent
            self.home_dir = QStandardPaths.standardLocations(QStandardPaths.StandardLocation.DesktopLocation)[0]
        
        def browse(self):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(None, 'Open File')
            except Exception as e:

                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        

        def open_database(self):
            self.database = self.fileName[0]
            self.assignments = pd.read_excel(self.database, sheet_name='Assigments')
            self.sales = pd.read_excel(self.database , sheet_name='Sales', index_col=0)
            self.associates = pd.read_excel(self.database , sheet_name='Associates', index_col=0)


        def extract_sales_list(self):
            self.open_database()
            active_assignments = self.assignments.loc[self.assignments['Status'] == "Active"]
            sales_list = active_assignments[['ID','Name', 'Client', 'Associate.ID','Client.ID','Location']].sort_values('Location')
            sales_list.rename(columns= {'ID':'Assignment ID'}, inplace=True)
            new_index = [self.sales.index[-1] + i for i in range(1, len(sales_list)+1)]
            sales_list.index = new_index
            sales_list = pd.concat([self.sales, sales_list], axis=0)
            ire_sales_list = sales_list.loc[sales_list['Location']== 'Ireland'].drop(columns='Location')
            bv_sales_list = sales_list.loc[sales_list['Location']== 'Netherlands'].drop(columns='Location')
            uk_sales_list = sales_list.loc[sales_list['Location']== 'United Kingdom'].drop(columns='Location')
            return ire_sales_list,  bv_sales_list, uk_sales_list
        
        def export_sales_list(self):
            ire, bv, uk = self.extract_sales_list()
            ire.to_excel(os.path.join(self.home_dir, 'CMIRE Sales List.xlsx'))
            bv.to_excel(os.path.join(self.home_dir, 'CMBV Sales List.xlsx'))
            uk.to_excel(os.path.join(self.home_dir, 'CMUK Sales List.xlsx'))  
        

    






            
        


        

